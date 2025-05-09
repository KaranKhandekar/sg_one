import os
import time
import pandas as pd
from datetime import datetime
from pathlib import Path
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from PIL import Image
import subprocess
import threading
import openpyxl.styles

class ImageProcessor(threading.Thread):
    def __init__(self, source_folder, num_designers, progress_callback, complete_callback, scan_callback):
        super().__init__()
        self.source_folder = source_folder
        self.num_designers = num_designers
        self.progress_callback = progress_callback
        self.complete_callback = complete_callback
        self.scan_callback = scan_callback
        self.supported_formats = {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff'}
        self.stats = {
            'total_images': 0,
            'white_background': 0,
            'non_white_background': 0,
            'extensions': {},
            'designer_files': {}
        }

    def run(self):
        try:
            start_time = time.time()
            
            # Create designer folders
            for i in range(self.num_designers):
                folder_path = os.path.join(self.source_folder, f'Designer_{i+1}')
                os.makedirs(folder_path, exist_ok=True)
                self.stats['designer_files'][f'Designer_{i+1}'] = []

            # First scan to count files and group by file ID
            image_groups = {}
            image_files = []
            for root, _, files in os.walk(self.source_folder):
                for file in files:
                    ext = os.path.splitext(file)[1].lower()
                    if ext in self.supported_formats:
                        file_id = self.extract_file_id(file)
                        if file_id:
                            if file_id not in image_groups:
                                image_groups[file_id] = []
                            image_groups[file_id].append(os.path.join(root, file))
                            image_files.append(os.path.join(root, file))
                            self.stats['extensions'][ext] = self.stats['extensions'].get(ext, 0) + 1
                            self.scan_callback(len(image_files))

            self.stats['total_images'] = len(image_files)
            processed = 0

            # Sort groups by size (number of files in each group)
            sorted_groups = sorted(image_groups.items(), key=lambda x: len(x[1]), reverse=True)
            
            # Initialize designer loads
            designer_loads = [0] * self.num_designers
            designer_groups = [[] for _ in range(self.num_designers)]
            
            # Distribute groups to designers
            for file_id, group_files in sorted_groups:
                # Find designer with minimum load
                min_load_designer = designer_loads.index(min(designer_loads))
                
                # Add group to the designer with minimum load
                designer_groups[min_load_designer].append((file_id, group_files))
                designer_loads[min_load_designer] += len(group_files)
            
            # Process the distributed groups
            for designer_index, groups in enumerate(designer_groups):
                designer_folder = os.path.join(self.source_folder, f'Designer_{designer_index + 1}')
                
                for file_id, group_files in groups:
                    for image_path in group_files:
                        image_file = os.path.basename(image_path)
                        dest_path = os.path.join(designer_folder, image_file)
                        self.stats['designer_files'][f'Designer_{designer_index + 1}'].append(image_file)

                        try:
                            os.rename(image_path, dest_path)
                            
                            if self.is_white_background(dest_path):
                                self.stats['white_background'] += 1
                                self.apply_mac_tag(dest_path, 6)
                            else:
                                self.stats['non_white_background'] += 1
                                self.apply_mac_tag(dest_path, 4)
                            
                            processed += 1
                            elapsed_time = time.time() - start_time
                            self.progress_callback(processed, self.format_time(elapsed_time), self.stats)

                        except Exception as e:
                            print(f"Error processing {image_file}: {str(e)}")

            self.create_excel_report(start_time)
            self.complete_callback(self.stats)

        except Exception as e:
            print(f"Error in processing thread: {str(e)}")

    def format_time(self, seconds):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        seconds = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def apply_mac_tag(self, file_path, tag_index):
        try:
            file_path = file_path.replace('"', '\\"')
            script = f"""osascript -e 'tell application "Finder" to set label index of (POSIX file "{file_path}" as alias) to {tag_index}'"""
            os.system(script)
        except Exception as e:
            print(f"Error applying tag: {str(e)}")

    def is_white_background(self, image_path):
        try:
            with Image.open(image_path) as img:
                width, height = img.size
                pixels = img.load()
                
                is_left_white = True
                for x in range(5):
                    for y in range(5):
                        if pixels[x, y][:3] != (255, 255, 255):
                            is_left_white = False
                            break
                    if not is_left_white:
                        break
                
                is_right_white = True
                for x in range(width-5, width):
                    for y in range(5):
                        if pixels[x, y][:3] != (255, 255, 255):
                            is_right_white = False
                            break
                    if not is_right_white:
                        break
                
                return is_left_white or is_right_white
        except Exception as e:
            print(f"Error checking white background: {e}")
            return False

    def extract_file_id(self, filename):
        if len(filename) >= 13 and filename[:13].isdigit():
            return filename[:13]
        elif len(filename) >= 12:
            return filename[:12]
        return None

    def create_excel_report(self, start_time):
        try:
            max_files = max(len(files) for files in self.stats['designer_files'].values()) if self.stats['designer_files'] else 0
            data = {designer: files + [''] * (max_files - len(files)) 
                   for designer, files in self.stats['designer_files'].items()}
            
            df = pd.DataFrame(data)
            excel_path = os.path.join(self.source_folder, 'SplitImg_Report.xlsx')
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Write the main data
                df.to_excel(writer, sheet_name='Designer Files', index=False)
                
                # Get the worksheet
                worksheet = writer.sheets['Designer Files']
                
                # Define colors
                dark_blue = "D9E1F2"  # Dark Blue, Text 2, Lighter 80%
                olive_green = "E2EFDA"  # Olive Green, Accent 3, Lighter 80%
                
                # Apply colors based on background detection
                for col in range(1, len(df.columns) + 1):
                    for row in range(2, len(df) + 2):  # Start from 2 to skip header
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value:
                            # Check if the file has white background
                            is_white = False
                            for designer, files in self.stats['designer_files'].items():
                                if cell_value in files:
                                    # Find the file in the source folder
                                    file_path = os.path.join(self.source_folder, designer, cell_value)
                                    if os.path.exists(file_path):
                                        is_white = self.is_white_background(file_path)
                                    break
                            
                            # Apply the appropriate color
                            fill_color = olive_green if is_white else dark_blue
                            worksheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
                                start_color=fill_color,
                                end_color=fill_color,
                                fill_type="solid"
                            )
                
                # Write the summary sheet
                extensions_text = ', '.join(f"{ext} ({count})" for ext, count in self.stats['extensions'].items())
                
                total_time = time.time() - start_time
                hours = int(total_time // 3600)
                minutes = int((total_time % 3600) // 60)
                seconds = int(total_time % 60)
                processing_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                
                summary_data = {
                    'Metric': [
                        'Total Images Processed',
                        'White Background Images',
                        'Non-White Background Images',
                        'Supported Extensions',
                        'Total Processing Time'
                    ],
                    'Value': [
                        self.stats['total_images'],
                        self.stats['white_background'],
                        self.stats['non_white_background'],
                        extensions_text,
                        processing_time
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            print(f"Excel report created: {excel_path}")
            return excel_path
            
        except Exception as e:
            print(f"Error creating Excel report: {str(e)}")
            return None

class SplitImageApp(ctk.CTkFrame):
    def __init__(self, root, dashboard):
        print("Initializing SplitImageApp...")  # Debug log
        super().__init__(root, fg_color="#1E1E1E")
        self.root = root
        self.dashboard = dashboard
        self.processor = None
        
        # Bind window state change event
        self.root.bind('<Unmap>', self.on_window_minimize)
        self.root.bind('<Map>', self.on_window_restore)
        
        # Create main container
        main_container = ctk.CTkFrame(self, fg_color="transparent")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Create left panel container with scroll
        left_container = ctk.CTkFrame(main_container, fg_color="transparent")
        left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Create scrollable frame for left panel
        self.scrollable_frame = ctk.CTkScrollableFrame(left_container, fg_color="transparent")
        self.scrollable_frame.pack(fill=tk.BOTH, expand=True)
        
        print("Creating left panel...")  # Debug log
        # Create left panel (Settings and Stats)
        left_panel = ctk.CTkFrame(self.scrollable_frame, fg_color="#252525", corner_radius=15)
        left_panel.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
        title_frame.pack(fill=tk.X, pady=20, padx=30)
        
        ctk.CTkLabel(
            title_frame,
            text="SG One Split Image",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color="#00F5C4"
        ).pack(side=tk.LEFT)
        
        # Description
        ctk.CTkLabel(
            left_panel,
            text="Split image files across multiple designer folders with automatic background detection",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF",
            wraplength=600
        ).pack(pady=(0, 20), padx=30)
        
        # Settings Frame
        settings_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
        settings_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 20))
        
        # Designer Configuration
        designer_frame = ctk.CTkFrame(settings_frame, fg_color="#333333", corner_radius=10)
        designer_frame.pack(fill=tk.X, pady=(0, 20))
        
        ctk.CTkLabel(
            designer_frame,
            text="Number of Designers (1-60):",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        ).pack(pady=(20, 10), padx=20)
        
        # Create a frame for the dropdown to control its width
        dropdown_container = ctk.CTkFrame(designer_frame, fg_color="transparent")
        dropdown_container.pack(fill=tk.X, pady=(0, 20), padx=20)
        
        # Create dropdown with scrollable values
        self.designer_input = ctk.CTkOptionMenu(
            dropdown_container,
            values=[str(i) for i in range(1, 61)],
            width=600,  # Increased width
            height=35,
            font=ctk.CTkFont(size=14),
            dropdown_font=ctk.CTkFont(size=14),
            command=self.validate_inputs,
            fg_color="#404040",
            button_color="#404040",
            button_hover_color="#505050",
            dropdown_fg_color="#404040",
            dropdown_hover_color="#505050"
        )
        self.designer_input.pack(fill=tk.X, expand=True)
        self.designer_input.set("Select number of designers")
        
        # Source Folder Selection
        folder_frame = ctk.CTkFrame(settings_frame, fg_color="#333333", corner_radius=10)
        folder_frame.pack(fill=tk.X, pady=(0, 20))
        
        ctk.CTkLabel(
            folder_frame,
            text="Source Folder:",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        ).pack(pady=(20, 10), padx=20)
        
        folder_input_frame = ctk.CTkFrame(folder_frame, fg_color="transparent")
        folder_input_frame.pack(fill=tk.X, pady=(0, 20), padx=20)
        
        self.folder_input = ctk.CTkEntry(
            folder_input_frame,
            placeholder_text="No folder selected",
            width=400,
            height=35
        )
        self.folder_input.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        browse_button = ctk.CTkButton(
            folder_input_frame,
            text="Browse...",
            width=100,
            height=35,
            command=self.browse_folder
        )
        browse_button.pack(side=tk.RIGHT)
        
        # Progress Frame
        progress_frame = ctk.CTkFrame(settings_frame, fg_color="#333333", corner_radius=10)
        progress_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.scan_progress_label = ctk.CTkLabel(
            progress_frame,
            text="Scanning Images...",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        )
        self.scan_progress_label.pack(pady=(20, 10), padx=20)
        
        self.scan_progress_bar = ctk.CTkProgressBar(
            progress_frame,
            width=400,
            height=8,
            corner_radius=4,
            progress_color="#00F5C4",
            fg_color="#404040"
        )
        self.scan_progress_bar.pack(pady=(0, 20), padx=20)
        self.scan_progress_bar.set(0)
        
        self.process_progress_label = ctk.CTkLabel(
            progress_frame,
            text="Processing Images...",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        )
        self.process_progress_label.pack(pady=(0, 10), padx=20)
        
        self.process_progress_bar = ctk.CTkProgressBar(
            progress_frame,
            width=400,
            height=8,
            corner_radius=4,
            progress_color="#00F5C4",
            fg_color="#404040"
        )
        self.process_progress_bar.pack(pady=(0, 20), padx=20)
        self.process_progress_bar.set(0)
        
        # Statistics Frame with two columns
        stats_frame = ctk.CTkFrame(settings_frame, fg_color="#333333", corner_radius=10)
        stats_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Create two columns for stats
        stats_columns = ctk.CTkFrame(stats_frame, fg_color="transparent")
        stats_columns.pack(fill=tk.X, padx=20, pady=20)
        
        # Column 1
        col1 = ctk.CTkFrame(stats_columns, fg_color="transparent")
        col1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.total_images_label = ctk.CTkLabel(
            col1,
            text="Total Images Processed: 0",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        )
        self.total_images_label.pack(pady=(0, 10))
        
        self.time_label = ctk.CTkLabel(
            col1,
            text="Time Taken: 00:00:00",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        )
        self.time_label.pack(pady=(0, 10))
        
        # Column 2
        col2 = ctk.CTkFrame(stats_columns, fg_color="transparent")
        col2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.white_bg_label = ctk.CTkLabel(
            col2,
            text="White Background Images: 0",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        )
        self.white_bg_label.pack(pady=(0, 10))
        
        self.non_white_bg_label = ctk.CTkLabel(
            col2,
            text="Non-White Background Images: 0",
            font=ctk.CTkFont(size=14),
            text_color="#FFFFFF"
        )
        self.non_white_bg_label.pack(pady=(0, 10))
        
        # Action Buttons
        buttons_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        buttons_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.run_button = ctk.CTkButton(
            buttons_frame,
            text="Start Processing",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40,
            command=self.start_processing,
            fg_color="#00F5C4",
            text_color="#000000",
            hover_color="#00D4A8",
            state="disabled"
        )
        self.run_button.pack(side=tk.LEFT, padx=(0, 10))
        
        reset_button = ctk.CTkButton(
            buttons_frame,
            text="Reset",
            font=ctk.CTkFont(size=14),
            height=40,
            command=self.reset_application,
            fg_color="#333333",
            hover_color="#404040"
        )
        reset_button.pack(side=tk.LEFT)
        
        # Create right panel (Logs) - Fixed position
        right_panel = ctk.CTkFrame(main_container, fg_color="#252525", corner_radius=15, width=300)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(10, 0), pady=0)
        right_panel.pack_propagate(False)
        
        # Log title
        log_title_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        log_title_frame.pack(fill=tk.X, pady=20, padx=20)
        
        ctk.CTkLabel(
            log_title_frame,
            text="Activity Log",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#00F5C4"
        ).pack(side=tk.LEFT)
        
        # Log text area
        self.log_text = scrolledtext.ScrolledText(
            right_panel,
            wrap=tk.WORD,
            font=("Courier", 12),
            bg="#333333",
            fg="#FFFFFF",
            insertbackground="#FFFFFF",
            relief="flat",
            height=20
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        self.log_text.configure(state='disabled')
        
        # Clear log button
        clear_button = ctk.CTkButton(
            right_panel,
            text="Clear Log",
            font=ctk.CTkFont(size=14),
            height=35,
            command=self.clear_log,
            fg_color="#333333",
            hover_color="#404040"
        )
        clear_button.pack(pady=(0, 20), padx=20)
        
        # Add initial log entry
        self.add_log("Split Image module initialized")
        
        # Bind resize event
        self.bind('<Configure>', self._on_resize)
    
    def _on_resize(self, event):
        # Update scrollable frame width when the frame is resized
        self.scrollable_frame.configure(width=event.width - 320)  # Account for right panel width
    
    def _on_mousewheel(self, event):
        self.scrollable_frame.yview_scroll(int(-1*(event.delta/120)), "units")
        
    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Image Folder")
        if folder:
            self.folder_input.delete(0, tk.END)
            self.folder_input.insert(0, folder)
            self.last_folder_path = folder
            self.validate_inputs()
            self.add_log(f"Source folder selected: {folder}")
    
    def validate_inputs(self, value=None):
        """Validate the input fields"""
        try:
            if value and value != "Select number of designers":
                num = int(value)
                if 1 <= num <= 60:
                    self.run_button.configure(state="normal")
                    return
            self.run_button.configure(state="disabled")
        except ValueError:
            self.run_button.configure(state="disabled")
    
    def start_processing(self):
        try:
            num_designers = int(self.designer_input.get())
            source_folder = self.folder_input.get()
            
            if 1 <= num_designers <= 60 and source_folder:
                self.processor = ImageProcessor(
                    source_folder,
                    num_designers,
                    self.update_progress,
                    self.processing_complete,
                    self.update_scan_progress
                )
                self.processor.start()
                self.run_button.configure(state="disabled")
                self.reset_progress()
                self.add_log(f"Starting processing with {num_designers} designers")
            else:
                self.add_log("Error: Please check your inputs")
        except ValueError:
            self.add_log("Error: Invalid number of designers")
        except Exception as e:
            self.add_log(f"Error: {str(e)}")
    
    def reset_progress(self):
        self.scan_progress_bar.set(0)
        self.process_progress_bar.set(0)
        self.total_images_label.configure(text="Total Images Processed: 0")
        self.white_bg_label.configure(text="White Background Images: 0")
        self.non_white_bg_label.configure(text="Non-White Background Images: 0")
        self.time_label.configure(text="Time Taken: 00:00:00")
    
    def update_scan_progress(self, count):
        self.scan_progress_bar.set(count)
        self.scan_progress_label.configure(text=f"Scanning Images... ({count} files found)")
        self.add_log(f"Found {count} image files")
    
    def update_progress(self, processed, time_taken, stats):
        total = stats['total_images']
        if total > 0:
            percentage = (processed / total)
            self.process_progress_bar.set(percentage)
        else:
            self.process_progress_bar.set(0)
            
        self.process_progress_label.configure(text=f"Processing Images... ({processed}/{total})")
        self.total_images_label.configure(text=f"Total Images Processed: {total}")
        self.white_bg_label.configure(text=f"White Background Images: {stats['white_background']}")
        self.non_white_bg_label.configure(text=f"Non-White Background Images: {stats['non_white_background']}")
        self.time_label.configure(text=f"Time Taken: {time_taken}")
    
    def processing_complete(self, stats):
        self.run_button.configure(state="normal")
        self.process_progress_bar.set(1)
        self.add_log("Processing complete! Excel report generated in the source folder.")
        
        report_path = os.path.join(self.folder_input.get(), 'SplitImg_Report.xlsx')
        messagebox.showinfo("Processing Complete", 
                          f"Successfully processed {stats['total_images']} images.\n\n"
                          f"White background: {stats['white_background']}\n"
                          f"Non-white background: {stats['non_white_background']}\n\n"
                          f"Report saved to:\n{report_path}")
    
    def reset_application(self):
        if messagebox.askyesno('Reset Split Image',
                             'Are you sure you want to reset? This will clear all inputs and progress.'):
            self.designer_input.set("Select number of designers")
            self.folder_input.delete(0, tk.END)
            self.reset_progress()
            self.scan_progress_label.configure(text="Scanning Images...")
            self.process_progress_label.configure(text="Processing Images...")
            self.run_button.configure(state="disabled")
            self.add_log("Application reset")
            
            # Restore last values
            if self.last_designer_count:
                self.designer_input.insert(0, self.last_designer_count)
            if self.last_folder_path:
                self.folder_input.insert(0, self.last_folder_path)
            self.validate_inputs()
    
    def add_log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state='normal')
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state='disabled')
    
    def clear_log(self):
        self.log_text.configure(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state='disabled')
        self.add_log("Log cleared")

    def on_window_minimize(self, event):
        """Handle window minimize event"""
        self.add_log("Window minimized")
        
    def on_window_restore(self, event):
        """Handle window restore event"""
        self.add_log("Window restored")
        self.root.deiconify()  # Ensure window is visible
        self.root.lift()  # Bring window to front
        self.root.focus_force()  # Force focus on window